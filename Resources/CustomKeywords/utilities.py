import email
import imaplib

from robot.api.deco import keyword
from openpyxl import load_workbook
from datetime import datetime, timedelta
import random
import pandas as pd
import string
from robot.libraries.BuiltIn import BuiltIn
from robot.api.deco import keyword

@keyword
def fetch_testdata_by_id(file_path, target_id):
    global workbook
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        header = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == target_id:
                data_dict = {}
                for col_name, value in zip(header, row):
                    if ',' in str(value):
                        data_dict[col_name] = [item.strip() for item in value.split(',')]
                    else:
                        data_dict[col_name] = value
                return data_dict

        # If the target_id is not found, you can raise an exception or return a specific value
        raise ValueError(f"Target ID '{target_id}' not found in the Excel file.")
    except Exception as e:
        print(f"Error reading Excel file: {e}")

@keyword
def calculate_new_date(input_day=13, month_offset=2):
    # Get the current date
    current_date = datetime.now()

    # Construct the input date using the current year and month, and the input day
    input_date = current_date.replace(day=input_day+1)

    # Calculate the new date
    new_date = input_date + timedelta(days=month_offset * 30)

    return new_date.strftime('%d/%m/%Y')


@keyword
def generate_random_name(length=4):
    vowels = 'aeiou'
    consonants = 'bcdfghjklmnpqrstvwxyz'
    name = ''
    for i in range(length):
        if i % 2 == 0:
            name += random.choice(consonants)
        else:
            name += random.choice(vowels)
    return name.capitalize()
@keyword
def random_assortment():
    if not hasattr(random_assortment, 'last_index'):
        random_assortment.last_index = 0

    built_in_names = ['WM10', 'WM11']
    if random_assortment.last_index == 0:
        name = 'WM10'
    else:
        name = 'WM11'
    random_assortment.last_index = (random_assortment.last_index + 1) % len(built_in_names)

    return name


@keyword
def generate_random_number():
    return random.randint(1, 2)

@keyword
def get_promo_names(file_path):
    promo_workbook = load_workbook(file_path)
    sheet = promo_workbook.active

    # Get header row
    header = [cell.value for cell in sheet[1]]

    # Find index of "promo_name" and "store_name" columns
    try:
        promo_name_index = header.index("promo_name")
    except ValueError:
        raise ValueError("Column 'promo_name' not found in the header.")

    try:
        store_name_index = header.index("store_name")
    except ValueError:
        raise ValueError("Column 'store_name' not found in the header.")

    # Create a dictionary to store promo names by store
    promo_by_store_dict = {}

    # Iterate over each row in the sheet
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from the second row
        promo_name = row[promo_name_index]
        store_name = row[store_name_index]

        # If store_name already exists in the dictionary, append promo_name to its value list
        if store_name in promo_by_store_dict:
            promo_by_store_dict[store_name].append(promo_name)
        # If store_name is new, create a new key-value pair
        else:
            promo_by_store_dict[store_name] = [promo_name]

    # Remove the entry None: [None] if it exists
    if None in promo_by_store_dict and len(promo_by_store_dict[None]) == 1 and promo_by_store_dict[None][0] is None:
        del promo_by_store_dict[None]
    promo_by_store_dict = {key: value for key, value in promo_by_store_dict.items() if value != [None]}
    return promo_by_store_dict

@keyword
def get_dictionary_value(dictionary, key_name):
    if key_name in dictionary:
        return dictionary[key_name]
    else:
        return None

@keyword
def fetch_promo_names(filename):
    try:
        # Read the Excel file
        df = pd.read_excel(filename)

        # Remove NaN values from 'promo_name' column
        promo_names = df['promo_name'].dropna().tolist()

        return promo_names
    except Exception as e:
        print(f"Error occurred: {e}")
        return None

# Promotions
@keyword
def get_formatted_date():
    # Get today's date
    today = datetime.today()
    # Format the date as 'April 2, 2024' without leading zero on the day
    formatted_date = today.strftime('%B %#d, %Y')
    return formatted_date

@keyword
def get_dob_fields(dob_str):
    try:
        dob_date_str = dob_str.split()[0]
        dob_date = datetime.strptime(dob_date_str, '%Y-%m-%d')
        day = dob_date.day
        month = dob_date.strftime('%B')
        year = dob_date.year
        return {'day': day, 'month': month, 'year': year}
    except ValueError:
        print("Invalid date format. Please provide date in the format MM/DD/YYYY.")
        return {'day': None, 'month': None, 'year': None}

streets = ['Main St', 'Oak St', 'Elm St', 'Maple Ave', 'Cedar Ln', 'Pine St']

@keyword
def generate_random_street_address():
    street_name = random.choice(streets)
    address = f"{street_name}"
    return address

@keyword
def generate_random_phone_number():
    # The first digit of a phone number should not be 0 or 1
    first_digit = random.randint(6, 9)
    # Generate the remaining 9 digits
    remaining_digits = [random.randint(0, 9) for _ in range(9)]
    # Combine all digits into a string
    phone_number = ''.join(map(str, [first_digit] + remaining_digits))
    return phone_number

@keyword
def generate_random_first_name():
    first_names = ['John', 'Emma', 'Michael', 'Olivia', 'William', 'Sophia', 'James', 'Ava', 'Alexander', 'Mia']
    return random.choice(first_names)

@keyword
def generate_random_last_name():
    last_name = ['sharma', 'watson', 'kali', 'jones', 'stark', 'newton', 'einstien', 'botsan']
    return random.choice(last_name)

@keyword
def generate_random_gender():
    genders = ['male', 'female']
    return random.choice(genders)

@keyword
def generate_random_email():
    domains = ['gmail.com', 'yahoo.com', 'hotmail.com', 'outlook.com', 'aol.com']
    username = ''.join(random.choices(string.ascii_letters + string.digits, k=random.randint(5, 10)))
    domain = random.choice(domains)
    return f"{username}@{domain}"

@keyword
def convert_item_list_to_dictionary(item_list):
    items_dict = {}

    # If item_list is not already a list, convert it into a single-element list
    if not isinstance(item_list, list):
        item_list = [item_list]

    # Iterate over each item in the items list
    for item in item_list:
        print("Current item:", item)  # Debugging print statement
        if isinstance(item, str):  # If the item is a string, treat it as a key-value pair
            key_value_pair = item.split(' : ')
            print("Split result:", key_value_pair)  # Debugging print statement
            if len(key_value_pair) != 2:
                print("Warning: Unexpected format for item:", item)
                continue
            key, value = key_value_pair
            value = value.strip()
            # Try to convert value to integer, otherwise keep as string
            try:
                value = int(value)
            except ValueError:
                pass
            items_dict[key.strip()] = value
        elif isinstance(item, list):  # If the item is a list, assume it's already in key-value pair format
            for item_str in item:
                key_value_pair = item_str.split(' : ')
                print("Split result:", key_value_pair)  # Debugging print statement
                if len(key_value_pair) != 2:
                    print("Warning: Unexpected format for item:", item_str)
                    continue
                key, value = key_value_pair
                value = value.strip()
                # Try to convert value to integer, otherwise keep as string
                try:
                    value = int(value)
                except ValueError:
                    pass
                items_dict[key.strip()] = value

    return items_dict

def remove_characters(data):
    # Define characters to remove
    chars_to_remove = ["-", ",", "₹", "+", "(", ")", " "]

    # Iterate over each character to remove
    for char in chars_to_remove:
        # Remove the character from the data
        data = data.replace(char, "")

    return data.strip()  # Remove leading and trailing whitespace

def convert_items_to_list(items):
    if isinstance(items, str):
        return [items]
    elif isinstance(items, list):
        return items
    else:
        raise ValueError("Unsupported type. Input should be either a string or a list.")

@keyword
def count_table_body_rows(locator):
    """
    Count the number of rows in a table body.

    Args:
    - locator: Locator of the table body element.

    Returns:
    - Integer: Number of rows in the table body.
    """
    selenium_lib = BuiltIn().get_library_instance('SeleniumLibrary')

    # Find the table body element
    table_body = selenium_lib.find_element(locator)

    # Find all rows within the table body
    rows = table_body.find_elements_by_tag_name("tr")

    return len(rows)

@keyword
def get_key_combination(keys):
    key_mapping = {
        "ctrl": "CONTROL",
        "alt": "ALT",
        "shift": "SHIFT",
        "cmd": "COMMAND",
        "win": "META"
    }

    keys_list = keys.lower().split('+')
    return '+'.join([key_mapping.get(key, key) for key in keys_list])

@keyword
def search_and_fetch_email(server, port, email_address, password, subject):
    try:
        # Connect to the email server
        mail = imaplib.IMAP4_SSL(server, port)
        print(f"Connecting to {server}:{port}")

        mail.login(email_address, password)
        print(f"Logged in as {email_address}")

        mail.select('inbox')
        print("Selected inbox")

        # Search for emails with the specified subject
        result, data = mail.search(None, f'(SUBJECT "{subject}")')
        if result != "OK":
            print(f"Search failed: {result}")
            return None, None
        email_ids = data[0].split()
        print(f"Email IDs found: {email_ids}")

        # Fetch the latest email and extract the body and attachments
        if email_ids:
            latest_email_id = email_ids[-1]
            result, data = mail.fetch(latest_email_id, '(RFC822)')
            if result != "OK":
                print(f"Fetch failed: {result}")
                return None, None
            raw_email = data[0][1]  # Get the raw email content

            # Parse the raw email into a message object
            email_message = email.message_from_bytes(raw_email)

            # Check if the subject contains the specified part
            if subject not in email_message['subject']:
                print(f"Subject mismatch: {email_message['subject']}")
                return None, None

            # Initialize email body variable and attachments list
            email_body = ""
            attachments = []

            # Function to recursively extract text from email parts and find attachments
            def extract_text_and_attachments(email_part):
                nonlocal email_body, attachments
                content_type = email_part.get_content_type()

                if content_type == "text/plain":
                    # Extract plain text body
                    email_body += email_part.get_payload(decode=True).decode()
                elif content_type == "text/html":
                    # Handle HTML content if needed
                    pass
                elif "image" in content_type:
                    # Skip image parts
                    pass
                elif content_type == "multipart/mixed":
                    # Handle mixed content (attachments)
                    for part in email_part.get_payload():
                        if part.get_content_maintype() == 'multipart':
                            # Recursively handle nested multipart
                            extract_text_and_attachments(part)
                        elif part.get_content_maintype() == 'application' and part.get('Content-Disposition'):
                            # This part is an attachment
                            attachments.append({
                                'filename': part.get_filename(),
                                'content_type': part.get_content_type(),
                                'content': part.get_payload(decode=True)
                            })
                        else:
                            # Handle other content types
                            extract_text_and_attachments(part)
                elif email_part.is_multipart():
                    # Recursively handle other multipart types
                    for part in email_part.get_payload():
                        extract_text_and_attachments(part)

            # Call the function with the root email message
            extract_text_and_attachments(email_message)

            # Print debug info
            print(f"Email body: {email_body[:100]}...")  # Print a snippet of the email body for debugging
            print(f"Attachments: {attachments}")  # Print attachments for debugging

            return email_body.strip(), attachments  # Return email body and attachments

        # If no email with the specified subject is found
        print("No emails found with the specified subject")
        return None, None
    except imaplib.IMAP4.error as e:
        print(f"IMAP error: {e}")
        return None, None

# Usage example
email_body, attachments = search_and_fetch_email(
    'imap.gmail.com', 993, 'dprimaryuser@gmail.com', 'aqwb qaoi bkwd jqbb', 'Your ${store_name} Invoice Receipt | ${cust_info_checkout.invoice_id}'
)

if email_body:
    print("\nEmail Body:")
    print(email_body)

if attachments:
    print("\nAttachments:")
    for attachment in attachments:
        print(f"Filename: {attachment['filename']}, Content type: {attachment['content_type']}")
else:
    print("\nNo attachments found.")

@keyword
def is_greater(first_number, second_number):
    return first_number > second_number

import json

@keyword
def extract_required_fields(json_data):
    # Load the JSON data
    data = json.loads(json_data)

    # Extract the required fields
    required_fields = {
        "api_token": data['sessions']['user']['api_token'],
        "v_id": data['sessions']['user']['v_id'],
        "store_id": data['sessions']['user']['store_id'],
        "licence": None,  # Assuming 'licence' is not present in the provided JSON
        "vu_id": data['sessions']['user']['vu_id'],
        "terminal_id": None,  # Assuming 'terminal_id' is not present in the provided JSON
        "url": "/revoke-licence?",
        "trans_from": "CLOUD_TAB_WEB",  # Based on the bearerToken 'source_platform'
        "session_id": data['sessions']['session']['id']
    }

    # Convert the result to JSON
    result_json = json.dumps(required_fields, indent=4)
    return result_json

@keyword
def extract_udidtoken(logs):
    try:
        # Check if logs is a list; if so, convert it to JSON string
        if isinstance(logs, list):
            logs = json.dumps(logs)

        # Parse the JSON string into a Python list
        logs_list = json.loads(logs)

        # Iterate through the list of log entries
        for entry in logs_list:
            if 'response' in entry:
                response_data = json.loads(entry['response'])
                if 'udidtoken' in response_data:
                    return response_data['udidtoken']

        return None
    except json.JSONDecodeError as e:
        print(f"Error decoding JSON: {e}")
        return None

@keyword
def merge_and_update_data_with_json_output(extracted_data, udidtoken, serial_key):
    # Ensure extracted_data is a dictionary, not a string
    if isinstance(extracted_data, str):
        extracted_data = json.loads(extracted_data)

    # Replace the udidtoken in extracted_data
    extracted_data['udidtoken'] = udidtoken

    # Attach the serial_key to the licence field
    extracted_data['licence'] = serial_key

    # Update the url and trans_from with static values
    extracted_data['url'] = "/revoke-licence?"
    extracted_data['trans_from'] = "CLOUD_TAB_WEB"

    # Convert the dictionary to a JSON string with double quotes
    updated_data_json = json.dumps(extracted_data, ensure_ascii=False)

    return updated_data_json

@keyword
def modify_json_data(json_data,user_mobile):
    # Ensure json_data is a dictionary, not a string
    if isinstance(json_data, str):
        json_data = json.loads(json_data)

    # Remove specified fields
    json_data.pop("terminal_id", None)
    json_data.pop("udidtoken", None)
    json_data.pop("url", None)

    # Add static data
    json_data["mobile"] = user_mobile
    json_data["url"] = "%2Fvendor%2Flogout"

    # Convert the dictionary to a JSON string with double quotes
    updated_data_json = json.dumps(json_data, ensure_ascii=False)

    return updated_data_json
